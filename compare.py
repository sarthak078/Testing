import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os
import logging

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('comparison.log'),
        logging.StreamHandler()
    ]
)

def log(message, level='info'):
    """Helper function for logging messages."""
    if level == 'debug':
        logging.debug(message)
    elif level == 'info':
        logging.info(message)
    elif level == 'warning':
        logging.warning(message)
    elif level == 'error':
        logging.error(message)
    else:
        logging.info(message)

# def find_most_recent_scan(domain_name, exclude_file=None, days_ago=None):
#     """
#     Find the most recent scan file for a given domain or a file from a specific number of days ago.
    
#     Parameters:
#     domain_name (str): Domain name to search for
#     exclude_file (str): Filename to exclude from the search
#     days_ago (int): Number of days ago for the old file (e.g., 7 for 7 days ago)
    
#     Returns:
#     str: Path to the most recent or specific scan file, or None if not found
#     """
#     today = datetime.datetime.now()
    
#     # If looking for an old file, check for the file created days_ago days ago
#     if days_ago:
#         target_date = today - datetime.timedelta(days=days_ago)
#         target_filename = f"{domain_name}_testing_{target_date.strftime('%Y-%m-%d')}.xlsx"
        
#         if os.path.exists(target_filename):
#             return target_filename  # Return the specific old file if found
#         else:
#             log(f"File from {days_ago} days ago ({target_filename}) not found. Falling back to the most recent file.", level='warning')
    
#     # If no old file is found or looking for the latest file, check for the newest file
#     scan_files = [
#         f for f in os.listdir('.')
#         if f.startswith(f"{domain_name}_testing_")  # Looking for testing pattern
#         and f.endswith('.xlsx')
#         and f != exclude_file
#     ]
    
#     # If no scan files are found, log and return None
#     if not scan_files:
#         log(f"No scan files found for domain {domain_name}.", level='error')
#         return None

#     # Return the most recent file based on modification time
#     most_recent_file = max(scan_files, key=lambda f: os.path.getmtime(f))
#     log(f"Most recent file for {domain_name}: {most_recent_file}", level='info')
#     return most_recent_file
import datetime
import os

def find_most_recent_scan(domain_name, exclude_file=None, days_ago=None):
    """
    Find the most recent scan file for a given domain or a file from a specific number of days ago.
    
    Parameters:
    domain_name (str): Domain name to search for.
    exclude_file (str): Filename to exclude from the search.
    days_ago (int): Number of days ago for the old file (e.g., 7 for 7 days ago).
    
    Returns:
    str: Path to the most recent or specific scan file, or None if not found.
    """
    today = datetime.datetime.now()
    
    # Step 1: If `days_ago` is specified, prioritize that file
    if days_ago is not None:
        target_date = today - datetime.timedelta(days=days_ago)
        target_filename = f"{domain_name}_testing_{target_date.strftime('%Y-%m-%d')}.xlsx"
        
        if os.path.exists(target_filename):
            log(f"File from {days_ago} days ago found: {target_filename}", level='info')
            return target_filename
        else:
            log(f"File from {days_ago} days ago ({target_filename}) not found. Falling back to the most recent file.", level='warning')
    
    # Step 2: Find all matching scan files, excluding the specified file
    scan_files = [
        f for f in os.listdir('.')
        if f.startswith(f"{domain_name}_testing_")  # Matches the naming pattern
        and f.endswith('.xlsx')
        and f != exclude_file
    ]
    
    if not scan_files:
        log(f"No scan files found for domain {domain_name}.", level='error')
        return None

    # Step 3: Find the most recent file by date in the filename (fallback mechanism)
    file_date_pairs = []
    for file in scan_files:
        try:
            # Extract date from filename
            date_str = file.split('_')[-1].replace('.xlsx', '')
            file_date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            file_date_pairs.append((file, file_date))
        except ValueError:
            log(f"Skipping file with invalid date format: {file}", level='warning')
    
    if not file_date_pairs:
        log(f"No valid scan files with dates found for domain {domain_name}.", level='error')
        return None

    # Sort files by date (descending) and get the latest
    most_recent_file = max(file_date_pairs, key=lambda x: x[1])[0]
    log(f"Most recent file for {domain_name}: {most_recent_file}", level='info')
    return most_recent_file




def compare_excel_reports(old_file, new_file, output_file):
    """
    Compare old and new Excel reports and generate a difference report.
    
    Parameters:
    old_file (str): Path to the old Excel file
    new_file (str): Path to the new Excel file
    output_file (str): Path to save the difference report
    """
    try:
        log(f"Starting comparison between {old_file} and {new_file}")
        
        # Load workbooks
        old_wb = openpyxl.load_workbook(old_file)
        new_wb = openpyxl.load_workbook(new_file)
        old_ws = old_wb.active
        new_ws = new_wb.active

        # Create dictionary of old data
        old_data = {}
        for row in range(7, old_ws.max_row + 1):  # Starting from row 7 as per original format
            subdomain = old_ws.cell(row=row, column=1).value
            if subdomain:
                old_data[subdomain] = {
                    'status': old_ws.cell(row=row, column=2).value,
                    'ports': old_ws.cell(row=row, column=3).value,
                }

        # Create dictionary of new data
        new_data = {}
        for row in range(7, new_ws.max_row + 1):
            subdomain = new_ws.cell(row=row, column=1).value
            if subdomain:
                new_data[subdomain] = {
                    'status': new_ws.cell(row=row, column=2).value,
                    'ports': new_ws.cell(row=row, column=3).value,
                }

        # Create difference workbook
        diff_wb = openpyxl.Workbook()
        diff_ws = diff_wb.active
        diff_ws.title = "New Findings"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        diff_ws.merge_cells('A1:E1')
        title_cell = diff_ws['A1']
        title_cell.value = "New Findings Report"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary section
        total_new = sum(1 for subdomain in new_data if subdomain not in old_data)
        total_changed = sum(1 for subdomain in new_data if subdomain in old_data and old_data[subdomain] != new_data[subdomain])
        
        diff_ws['A3'] = "Comparison Date:"
        diff_ws['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        diff_ws['A4'] = "Previous Scan File:"
        diff_ws['B4'] = old_file
        diff_ws['A5'] = "New Scan File:"
        diff_ws['B5'] = new_file
        diff_ws['D3'] = "New Subdomains Found:"
        diff_ws['E3'] = total_new
        diff_ws['D4'] = "Changed Subdomains:"
        diff_ws['E4'] = total_changed

        # Headers
        headers = ["Subdomain", "Status", "Open Ports", "Change Type", "Details"]
        for col, header in enumerate(headers, start=1):
            cell = diff_ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Compare and write differences
        row = 8
        for subdomain, new_info in new_data.items():
            if subdomain not in old_data:
                # New subdomain found
                diff_ws.cell(row=row, column=1, value=subdomain).border = border
                diff_ws.cell(row=row, column=2, value=new_info['status']).border = border
                diff_ws.cell(row=row, column=3, value=new_info['ports']).border = border
                diff_ws.cell(row=row, column=4, value="New Subdomain").border = border
                diff_ws.cell(row=row, column=5, value="Newly discovered subdomain").border = border
                row += 1
            elif old_data[subdomain] != new_info:
                # Changed subdomain
                old_info = old_data[subdomain]
                changes = []
                if old_info['status'] != new_info['status']:
                    changes.append(f"Status changed: {old_info['status']} → {new_info['status']}")
                if old_info['ports'] != new_info['ports']:
                    changes.append(f"Ports changed: {old_info['ports']} → {new_info['ports']}")

                if changes:
                    diff_ws.cell(row=row, column=1, value=subdomain).border = border
                    diff_ws.cell(row=row, column=2, value=new_info['status']).border = border
                    diff_ws.cell(row=row, column=3, value=new_info['ports']).border = border
                    diff_ws.cell(row=row, column=4, value="Changed").border = border
                    diff_ws.cell(row=row, column=5, value="\n".join(changes)).border = border
                    row += 1

        # Auto-adjust column widths
        for column in diff_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            diff_ws.column_dimensions[column_letter].width = adjusted_width

        # Save the difference report
        diff_wb.save(output_file)
        log(f"Difference report generated successfully: {output_file}")
        
        # Return summary statistics
        return {
            'new_subdomains': total_new,
            'changed_subdomains': total_changed,
            'total_differences': total_new + total_changed
        }
        
    except Exception as e:
        log(f"Error generating difference report: {e}", level='error')
        raise

if __name__ == "__main__":
    # Example usage
    domain = "nishantbanjade.com.np"  # Replace with your domain
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    new_file = f"{domain}_testing_{current_date}.xlsx"
    
    # Find most recent old report
    old_file = find_most_recent_scan(domain, exclude_file=new_file, days_ago=7)

    
    if old_file:
        diff_file = f"{domain}_differences_{current_date}.xlsx"
        log(f"Comparing {old_file} with {new_file}")
        
        # Generate difference report
        stats = compare_excel_reports(old_file, new_file, diff_file)
        
        # Print summary
        print("\nComparison Summary:")
        print(f"New subdomains found: {stats['new_subdomains']}")
        print(f"Changed subdomains: {stats['changed_subdomains']}")
        print(f"Total differences: {stats['total_differences']}")
        print(f"Difference report saved to: {diff_file}")
    else:
        log("No previous scan files found. Cannot generate comparison report.")