import requests
import subprocess
import concurrent.futures
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import dns.resolver
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import ssl
import OpenSSL
from datetime import datetime
import time
import logging
import os
import datetime


SUBLIST3R_PATH=r"C:\Users\spokhrel\AppData\Local\Programs\Python\Python311\Scripts\sublist3r.exe"
NMAP_PATH = r"C:\Program Files (x86)\Nmap\nmap.exe" 
AMASS_PATH = r"C:\amass_Windows_amd64\amass.exe" 
ASSETFINDER_PATH = r"C:\\Users\\spokhrel\\go\\bin\\assetfinder.exe"
SUBFINDER_PATH = r"C:\\Users\\spokhrel\\go\\bin\\subfinder.exe"


# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('domain.log'),
        logging.StreamHandler()
    ]
)

def log(message, level='info'):
    if level == 'debug':
        logging.debug(message)
    elif level == 'info':
        logging.info(message)
    elif level == 'warning':
        logging.warning(message)
    elif level == 'error':
        logging.error(message)
    else:
        logging.info(message)


# Define your API keys for services here if needed
def google_dorking(domain):
    log("Fetching subdomains from dork started")
    """Use Google Dorking to find subdomains."""
    subdomains = set()
    log("googledorking started")
    query = f"site:{domain} -www"
    for start in range(0, 100, 10):  # Retrieve multiple pages
        url = f"https://www.google.com/search?q={query}&start={start}"
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            for link in soup.find_all('a', href=True):
                parsed_url = urlparse(link['href'])
                netloc = parsed_url.netloc
                if netloc.endswith(domain):
                    subdomains.add(netloc)
    return subdomains

def get_subdomains_from_crt_sh(domain):
    log("Fetching subdomains from DNS started")
    """Fetch subdomains using crt.sh."""
    
    url = f"https://crt.sh/?q=%.{domain}&output=json"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        subdomains = set()
        for entry in data:
            names = entry['name_value'].split('\n')
            for name in names:
                if name.endswith(domain) and name != domain:
                    subdomains.add(name)
        return list(subdomains)
    return []

def get_subdomains_from_dns(domain):
    log("Fetching subdomains from DNS started")
    subdomains = set()
    try:
        answers = dns.resolver.resolve(domain, 'A')
        subdomains.update([str(rdata) + '.' + domain for rdata in answers])
        
        answers = dns.resolver.resolve(domain, 'AAAA')
        subdomains.update([str(rdata) + '.' + domain for rdata in answers])
        
        answers = dns.resolver.resolve(domain, 'CNAME')
        subdomains.update([str(rdata.target) for rdata in answers])
        
        answers = dns.resolver.resolve(domain, 'MX')
        subdomains.update([str(rdata.exchange) + '.' + domain for rdata in answers])
    except Exception as e:
        log(f"Error resolving DNS records: {e}")
    return subdomains

def get_subdomains_from_web_crawl(domain):
    log("Fetching subdomains from web_crawl started")
    subdomains = set()
    try:
        response = requests.get(f"https://{domain}", timeout=20, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36'
        }, allow_redirects=True)
        log(f"Final URL after redirects: {response.url}")
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            for link in soup.find_all('a', href=True):
                parsed_url = urlparse(link['href'])
                if parsed_url.netloc.endswith(domain):
                    subdomains.add(parsed_url.netloc)
        else:
            log(f"Received non-200 status code: {response.status_code}")
    except Exception as e:
        log(f"Error crawling website: {e}")
    return subdomains


def get_subdomains_from_cert(domain):
    log("Fetching subdomains from cert started")
    subdomains = set()
    try:
        cert = ssl.get_server_certificate((domain, 443))
        x509 = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, cert)
        for i in range(x509.get_extension_count()):
            ext = x509.get_extension(i)
            if ext.get_short_name().decode() == 'subjectAltName':
                subdomains.update([name.split(':')[1] for name in ext.__str__().split(',') if 'DNS:' in name and domain in name])
    except Exception as e:
        log(f"Error fetching certificate: {e}")
    return subdomains

# def get_subdomains_from_sublist3r(domain):
#     """Fetch subdomains using Sublist3r."""
#     log("sublist3r started")
#     subdomains = set()
#     try:
#         output_file = 'sublist3r_output.txt'
#         # Run Sublist3r command
#         result = subprocess.run([SUBLIST3R_PATH,'sublist3r', '-d', domain, '-o', output_file], capture_output=True, text=True)
        
#         # Check if Sublist3r ran successfully
#         if result.returncode == 0:
#             with open(output_file, 'r') as file:
#                 for line in file:
#                     subdomains.add(line.strip())
#         else:
#             log(f"Sublist3r returned an error: {result.stderr}")
#     except Exception as e:
#         log(f"Error running Sublist3r: {e}")
#     return list(subdomains)
def get_subdomains_from_sublist3r(domain):
    """Fetch subdomains using Sublist3r."""
    log("sublist3r started")
    
    subdomains = set()
    try:
        # Get the absolute path for output file in the current working directory
        output_file = os.path.join(os.getcwd(), 'sublist3r_output.txt')

        log(f"Current working directory: {os.getcwd()}")

        # Assuming SUBLIST3R_PATH is defined correctly
          # Update this with your actual Sublist3r path
        
        # Run Sublist3r command
        log(f"Running Sublist3r command for domain: {domain}")
        result = subprocess.run(['sublist3r', '-d', domain, '-o', output_file], capture_output=True, text=True)
        
        # Check if Sublist3r ran successfully
        if result.returncode == 0:
            log(f"Sublist3r completed successfully. Reading the output file: {output_file}")
            with open(output_file, 'r') as file:
                for line in file:
                    subdomains.add(line.strip())
        else:
            log(f"Sublist3r returned an error: {result.stderr}", level='error')
            log(f"Sublist3r stdout: {result.stdout}")
    except Exception as e:
        log(f"Error running Sublist3r: {e}", level='error')
    
    log(f"Found {len(subdomains)} subdomains.")
    return list(subdomains)












# def get_subdomains_from_sublist3r(domain):
#     """Fetch subdomains using Sublist3r."""
#     log("Sublist3r started")
#     subdomains = set()
    
#     try:
#         # Set the output file to an absolute path to avoid working directory issues
#         output_file = os.path.join(os.getcwd(), 'sublist3r_output.txt')
        
#         # Construct the full Sublist3r command
#         sublist3r_command = [SUBLIST3R_PATH, 'sublist3r', '-d', domain, '-o', output_file]
        
#         # Log the command for debugging purposes
#         log(f"Running command: {' '.join(sublist3r_command)}")
        
#         # Run the Sublist3r command
#         result = subprocess.run(sublist3r_command, capture_output=True, text=True)
        
#         # Check if Sublist3r ran successfully
#         if result.returncode == 0:
#             # Wait for the output file to be created
#             if os.path.exists(output_file):
#                 with open(output_file, 'r') as file:
#                     for line in file:
#                         subdomains.add(line.strip())
#                 log(f"Found {len(subdomains)} subdomains.")
#             else:
#                 log(f"Error: Output file '{output_file}' not found.")
#         else:
#             log(f"Sublist3r returned an error: {result.stderr}")
#             log(f"Sublist3r output: {result.stdout}")
#     except subprocess.CalledProcessError as e:
#         log(f"Sublist3r command failed with error: {e}")
#     except Exception as e:
#         log(f"Error running Sublist3r: {e}")
    
#     return list(subdomains)
# def get_subdomains_from_securitytrails(domain):
#     log("Fetching subdomains from securitytrials started")
#     headers = {"APIKEY": "XirNf7pFYI3m8oB7NqWxMzXfogWD1gge"}  # Add your API key here
#     url = f"https://api.securitytrails.com/v1/domain/{domain}/subdomains"
#     response = requests.get(url, headers=headers)
#     if response.status_code == 200:
#         data = response.json()
#         return {f"{sub}.{domain}" for sub in data['subdomains']}
#     return set()

# def get_subdomains_from_shodan(domain):
#     log("Fetching subdomains from shodan started")
#     url = f"https://api.shodan.io/dns/domain/{domain}?key=JEyVGZt7LNsRRlB55CxfyNG0E59I208p"  # Add your API key here
#     response = requests.get(url)
#     if response.status_code == 200:
#         return {f"{sub}.{domain}" for sub in response.json()['subdomains']}
#     return set()

# def get_subdomains_from_virustotal(domain):
#     log("Fetching subdomains from virus started")
#     headers = {"x-apikey": "912171d5e26b548f2dee8fd0462b2e4a7286f83b28bfbca4d8b27f9e68288865"}  # Add your API key here
#     url = f"https://www.virustotal.com/api/v3/domains/{domain}/subdomains"
#     response = requests.get(url, headers=headers)
#     if response.status_code == 200:
#         return {item['id'] for item in response.json()['data']}
#     return set()

# def get_subdomains_from_amass(domain):
#     """Fetch subdomains using Amass."""
#     log("Amass started")
#     subdomains = set()
#     try:
#         output_file = 'amass_output.txt'
#         # Run Amass command
#         result = subprocess.run(['amass', 'enum', '-d', domain, '-o', output_file], capture_output=True, text=True)

#         # Check if Amass ran successfully
#         if result.returncode == 0:
#             with open(output_file, 'r') as file:
#                 for line in file:
#                     subdomains.add(line.strip())
#         else:
#             log(f"Amass returned an error: {result.stderr}")
#     except Exception as e:
#         log(f"Error running Amass: {e}")
#     return list(subdomains)

def clean_up_output_files():
    """Delete the output files of tools after Excel is generated."""
    output_files = [
        'sublist3r_output.txt',
        'amass_output.txt',
        'subfinder_output.txt',
        'assetfinder_output.txt',
        'sublist3r_output.txt'  # Add any other files you wish to delete
    ]
    for file in output_files:
        if os.path.exists(file):
            try:
                os.remove(file)
                log(f"Deleted file: {file}")
            except Exception as e:
                log(f"Error deleting {file}: {e}", level='error')

def get_subdomains_from_amass(domain):
    """Fetch subdomains using Amass."""
    log("Amass started")
    subdomains = set()
    
    try:
        # Set the output file to an absolute path to avoid working directory issues
        output_file = os.path.join(os.getcwd(), 'amass_output.txt')
        
        # Construct the full Amass command
        amass_command = ['amass', 'enum', '-d', domain, '-o', output_file]
        
        # Log the command for debugging purposes
        log(f"Running command: {' '.join(amass_command)}")
        
        # Run the Amass command
        result = subprocess.run(amass_command, capture_output=True, text=True)
        
        # Log the stderr and stdout to understand what's happening
        if result.returncode != 0:
            log(f"Amass returned an error: {result.stderr}")
            log(f"Amass output: {result.stdout}")
        else:
            log(f"Amass completed successfully. Checking output file...")
            # Check if the output file exists and is non-empty
            if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                with open(output_file, 'r') as file:
                    for line in file:
                        subdomains.add(line.strip())
                log(f"Found {len(subdomains)} subdomains.")
            else:
                log("Error: Output file is empty or not created.")
                
    except subprocess.CalledProcessError as e:
        log(f"Amass command failed with error: {e}")
    except Exception as e:
        log(f"Error running Amass: {e}")
    
    return list(subdomains)

def get_subdomains_from_subfinder(domain):
    """Fetch subdomains using Subfinder."""
    log("Subfinder started")
    subdomains = set()
    try:
        output_file = 'subfinder_output.txt'
        # Run Subfinder command
        result = subprocess.run(['subfinder', '-d', domain, '-o', output_file], capture_output=True, text=True)

        # Check if Subfinder ran successfully
        if result.returncode == 0:
            with open(output_file, 'r') as file:
                for line in file:
                    subdomains.add(line.strip())
        else:
            log(f"Subfinder returned an error: {result.stderr}")
    except Exception as e:
        log(f"Error running Subfinder: {e}")
    return list(subdomains)

def get_subdomains_from_assetfinder(domain):
    """Fetch subdomains using Asset Finder."""
    log("Asset Finder started")
    subdomains = set()
    try:
        # Run Asset Finder command and capture output
        output_file = 'assetfinder_output.txt'  # Define the output file name
        result = subprocess.run(['assetfinder',domain], capture_output=True, text=True)

        # Check if Asset Finder ran successfully
        if result.returncode == 0:
            # Save the found subdomains to output file
            with open(output_file, 'w') as file:
                for line in result.stdout.splitlines():
                    subdomains.add(line.strip())  # Collect subdomains from stdout
                    file.write(f"{line.strip()}\n")  # Write each subdomain to the file
            log(f"Found and saved {len(subdomains)} subdomains to {output_file}.")
        else:
            log(f"Asset Finder returned an error: {result.stderr}", level='error')
    except Exception as e:
        log(f"Error running Asset Finder: {e}", level='error')
    
    return list(subdomains)




def get_all_subdomains(domain):
    """Combine subdomains from all sources."""
    all_subdomains = set()
    all_subdomains.update(get_subdomains_from_crt_sh(domain))
    all_subdomains.update(get_subdomains_from_dns(domain))
    all_subdomains.update(get_subdomains_from_web_crawl(domain))
    all_subdomains.update(get_subdomains_from_cert(domain))
    all_subdomains.update(get_subdomains_from_sublist3r(domain))
    # all_subdomains.update(get_subdomains_from_securitytrails(domain))
    # all_subdomains.update(get_subdomains_from_shodan(domain))
    # all_subdomains.update(get_subdomains_from_virustotal(domain))
    # all_subdomains.update(get_subdomains_from_amass(domain))
    all_subdomains.update(google_dorking(domain)) 
    all_subdomains.update(get_subdomains_from_subfinder(domain))
    all_subdomains.update(get_subdomains_from_assetfinder(domain))
    return list(all_subdomains)

def run_nmap_scan(subdomain):
    """Run Nmap scan on a subdomain."""
    open_ports = []
    try:
        # Add additional ports to the Nmap command
        command = [
            NMAP_PATH, 
            '-p', '21,22,25,53,80,110,143,443,3306,5432,6379,27017,8080,8443,5000,9000', 
            '--open', 
            subdomain
        ]
        result = subprocess.run(command, capture_output=True, text=True)

        # Parse Nmap output to extract open ports
        for line in result.stdout.splitlines():
            if '/tcp' in line and 'open' in line:
                port = line.split('/')[0].strip()
                open_ports.append(port)
    except Exception as e:
        print(f"Error running Nmap on {subdomain}: {e}")
    
    time.sleep(5) # Increased delay to avoid overwhelming the network
    return subdomain, open_ports

def scan_subdomains(domain):
    subdomains = get_all_subdomains(domain)
    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        future_to_subdomain = {executor.submit(run_nmap_scan, sub): sub for sub in subdomains}
        for future in concurrent.futures.as_completed(future_to_subdomain):
            try:
                results.append(future.result())
            except Exception as exc:
                log(f'Error during scan: {exc}')
    return results


def generate_excel_report(results, filename):
    """Generate an Excel report with scan results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subdomain Scan Results"

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "Subdomain Scan Results"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Scan info
    ws['A3'] = "Scan Date:"
    ws['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A4'] = "Total Subdomains Scanned:"
    ws['B4'] = len(results)

    # Headers
    headers = ["Subdomain", "Status", "Open Ports", "Recommendation"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data
    row = 7
    for subdomain, open_ports in results:
        if not open_ports:  # Inactive subdomains
            ws.cell(row=row, column=1, value=subdomain).border = border
            ws.cell(row=row, column=2, value="Inactive").border = border
            ws.cell(row=row, column=3, value=", ".join(map(str, open_ports))).border = border
            ws.cell(row=row, column=4, value="Consider deletion").border = border
        else:  # Active subdomains
            ws.cell(row=row, column=1, value=subdomain).border = border
            ws.cell(row=row, column=2, value="Active").border = border
            ws.cell(row=row, column=3, value=", ".join(map(str, open_ports))).border = border
            ws.cell(row=row, column=4, value="Monitor regularly").border = border
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)

# Example usage
clean_up_output_files()

if __name__ == "__main__":
    domain = "nishantbanjade.com.np"
    log(f"Starting comprehensive scan of {domain} subdomains...")
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"{domain}_testing_{current_date}.xlsx"
    scan_results = scan_subdomains(domain)
    generate_excel_report(scan_results, filename)
    clean_up_output_files()